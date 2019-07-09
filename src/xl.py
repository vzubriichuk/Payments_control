# -*- coding: utf-8 -*-
"""
Created on Fri Jul  5 12:24:40 2019

@author: v.shkaberda
"""
from decimal import Decimal
from log_error import writelog
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from openpyxl.utils import get_column_letter
from os import path
import time

def export_to_excel(headers, rows):
    try:
        wb = Workbook()
        # Grab the active worksheet
        ws = wb.active
        # Headers
        ws.append(key for key in headers.keys())
        # Rows
        for row in rows:
            ws.append(tuple(row))
        # Customizing style
        for col, width in enumerate(headers.values()):
            cell = ws.cell(row=1, column=col+1)
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(border_style='thin',
                                             color='FF000000'))
            cell.fill = PatternFill("solid", fgColor="D3D3D3")
            ws.column_dimensions[get_column_letter(col+1)].width = width / 5
        # Save the file
        time_txt = time.strftime("%d-%m-%Y %H.%M.%S", time.localtime())
        wb.save(path.join(Path.home(), 'Desktop',
                          'Заявки на платежи ({}).xlsx'.format(time_txt)))
        return 1
    except Exception as e:
        writelog(e)
        print(e)


if __name__ == '__main__':
    export_to_excel(headers={'abc':50, 'def':30}, rows=((2, 42), ('a', 'b')))